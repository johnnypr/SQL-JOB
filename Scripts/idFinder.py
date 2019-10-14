import xlrd
import numpy as np
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import time
from openpyxl.styles import *

class idFinder: 
    book = Workbook()
    arrSheet=[]
    s1 = []
    s2 = []
    s3 = []
    
    
    def __init__(self):

        fileName = None
        currentRow =0
        currentColumn = 0
    
    def dataSets(self,file1,file2):
        loc1 = (file1)
        loc2 = (file2)

        fileToLink = xlrd.open_workbook(loc1) 
        confirmID = xlrd.open_workbook(loc2)

        sheet1 = fileToLink.sheet_by_index(0)
        sheet2 = confirmID.sheet_by_index(0) 
 
        j=1
        i=1
        
        for i in range(1,sheet2.nrows):
            self.s1.append(str((sheet2.cell(i,0).value)))
        
        
        # print(self.s2)

        for j in range(1,sheet2.nrows):
            self.s2.append(str((sheet2.cell(j,1).value)))
        
        # print(self.s1)

        for n in range(sheet1.nrows):
            if sheet1.cell(n,0).value is not '':
                temp = str((sheet1.cell(n,0).value))
                self.s3.append(temp)

        print("Matches =", len(self.s3))
        
      
    
    def setFileName(self,fileName):
        if ".xlsx" not in fileName:
            self.fileName = str(fileName)+".xlsx"
            
        self.fileName = str(self.fileName)
        print(self.fileName)
    
    def createSheet(self,name):
        sheet = self.book.active
        
        self.arrSheet.append(sheet)

    def save(self):
        self.book.save(self.fileName)

    def createBase(self,r=None,sheet_num=None):
        index = 0
        data = self.s1
        data2 = self.s2
        data3 = self.s3
        
        yellow = colors.Color(rgb='FFFF00')
        
        red = colors.Color(rgb='ff0000')
        
        blue = colors.Color(rgb='2200ff')
        
    
        if r == None:
            r = len(data)

        if sheet_num != None:
            index = sheet_num

        sheet = self.arrSheet[index]
        sheet.title = "Test2"
                
        for i in range(r):
            cellref = sheet.cell(row=i+1,column=1)
            cellref.value = data2[i]
            cellref = sheet.cell(row=i+1,column=2)
            cellref.value = data[i]
        
        for j in range(len(data)):
            cellref = sheet.cell(row=j+1,column=2)
            if cellref.value in data3:
                my_fill = fills.PatternFill(patternType='solid', fgColor=yellow)
                cellref.fill = my_fill
            else:
                my_fill = fills.PatternFill(patternType='solid', fgColor=red)
                cellref.fill = my_fill

    


    

        
             


    