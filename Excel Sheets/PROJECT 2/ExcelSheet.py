import xlrd
from os import system,name
import numpy as np
from termcolor import colored
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from time import sleep

class ExcelSheet: 
    book = Workbook()
    arrSheet=[]
    s1 = []
    s2 = [] 
    f1 = None
    f2 = None
    key1 = []
    key2 = []
    usrKey1 = []
    usrKey2 = []
    sheet1 = None
    sheet2 = None
    def __init__(self):
        fileName = None
        self.setFileName()
        self.createSheet()
        

    def dataSets(self,file1,file2):
        #Sets File Name class variables
        self.f1 = file1
        self.f2 = file2

        #Initiates the Connection from The Excel Sheet to the Python Scripts
        excel_one = xlrd.open_workbook(file1) 
        excel_two = xlrd.open_workbook(file2)

        #In case there are more than one sheet in the Excel file we can change it manually by switching the index
        self.sheet1 = excel_one.sheet_by_index(0)
        self.sheet2 = excel_two.sheet_by_index(0) 
 
        #Load the Excel column keys so the user may pick which columns they may want to load                
        for i in range(self.sheet1.ncols):
            self.key1.append(str((self.sheet1.cell(0,i).value).upper()))

        for j in range(self.sheet2.ncols):
            self.key2.append(str((self.sheet2.cell(0,j).value).upper()))
        
    def chooseData(self):
        #Stores users decisions in arrays
        choice_one =[]
        choice_two =[]

        #CHANGE FROM INDEX TO NAME
        #Ask user to choose which data to load
        print("From file, ",colored(self.f1,'green')," choose which colums to load:  ")
        print(self.key1)
        usrInput = input("1. Choose the keys you which to load from the above list, to move to the next list enter \"x\":  ")
        while usrInput != "x":
            if usrInput != 'x':
                choice_one.append(usrInput)
            usrInput = input('Add another or enter \"x\" to quit: ')
        #print(choice_one)
        
        print("From file ",colored(self.f2,'green')," choose which columns to load")
        print(self.key2)
        usrInput = input("1. Choose the keys you which to load from the above list, to move to the next list enter \"x\":  ")
        while usrInput != "x":
            if usrInput != 'x':
                choice_two.append(usrInput)
            usrInput = input("Add another or enter \"x\" to quit: ")
        #print(choice_two)

        print('----------------------------------------------------------')

        print("From ", colored(self.f1,'green'), " ,we loaded","\n", choice_one)
        print("From ", colored(self.f2,'green'), " ,we loaded","\n", choice_two)
        print('----------------------------------------------------------')

    
        usrInput= input("Do you wish to restart the choice? [y]/[n]: ")

        print('----------------------------------------------------------')
        
        if usrInput == 'y':
            choice_one = []
            choice_two = []
            self.clear()
            self.chooseData()
        elif usrInput == 'n':
            self.clear()
            self.parseData(choice_one,choice_two)

    def parseData(self,choice_one,choice_two):
        for i in range(len(choice_one)):
            self.usrKey1.append(self.key1[int(choice_one[i])])

        for i in range(len(choice_two)):
            self.usrKey2.append(self.key2[int(choice_two[i])])
    
        
        if len(self.usrKey1)>1:
            print(self.usrKey1)
            usrInput = input("From file " + str(colored(self.f1,'green')) + " which will be used to compare? ")
            x = self.usrKey1[int(usrInput)]
        else:
            x = self.usrKey1[0]
        
            
        if len(self.usrKey2)>1:
            print(self.usrKey2)
            usrInput = input("From file "+ str(colored(self.f2,'green')) +" which will be used to compare? ")
            y = self.usrKey2[int(usrInput)]
        else:
            y = self.usrKey2[0]
        
        print('----------------------------------------------------------')


        print(colored("These will be the ID's we will be comparing",'green'))
        print("X: ",x)
        print("Y: ",y)
        print('----------------------------------------------------------')


       
        for i in range(1,self.sheet1.nrows):
            self.s1.append(str((self.sheet1.cell(i,self.findColumn(self.key1,x)).value)))

        for j in range(1,self.sheet2.nrows):
            self.s2.append(str((self.sheet2.cell(j,self.findColumn(self.key2,y)).value)))

        self.compare()

    def findColumn(self,key,id):
       for i in range(len(key)):
           if key[i] == id:
               return i
       
    def setFileName(self):
        fileName = input("Enter a filename: ")
        if ".txt" not in fileName:
            self.fileName = str(fileName)+".txt"
            
        self.fileName = str(self.fileName)
        print("Filename: ", colored(self.fileName, 'green'))
        stdIn = input("Is this correct? [y]/[n] ")

        if stdIn != 'y':
            self.fileName =None
            self.clear()
            self.setFileName()
        
        self.clear()
    
    def createSheet(self):
        sheet = self.book.active
        self.arrSheet.append(sheet)

    def save(self):
        self.book.save(self.fileName)

    def createBase(self,r=None,sheet_num=None):
        index = 0
        data = self.s1
        
  
        if r == None:
            r = len(data)-1

        if sheet_num != None:
            index = sheet_num

        sheet = self.arrSheet[index]
        sheet.title = "Test"


        
        for i in range(1,r):
            cellref = sheet.cell(row=i,column=2)
            cellref.value = data[i].upper()
    
        # Print iterations progress
         
    def clear(self): 
        # for windows 
        if name == 'nt': 
            _ = system('cls') 
  
        # for mac and linux(here, os.name is 'posix') 
        else: 
            _ = system('clear') 

    def compare(self,s=0):
        percent = 0
        sheet = self.arrSheet[0]

        x  = None
        y = None
        
        file = open(self.fileName,"w+") 

        for i in tqdm(range(s,len(self.s1))):
            x = str(self.s1[i])
            for j in range(len(self.s2)):
                y = str(self.s2[j])
                if x[0] != y[0]:
                    continue
                else:
                    try:
                        c1 = x.split()
                        c2 = y.split()
                        c1 = c1[0] + " " + c1[1]
                        c2 = c2[0] + " " + c2[1]
                    except:
                        c1 = x
                        c2 = y
                    finally:
                        if c1 == c2:
                            result = "X: " +str(x) +" Y: "+ str(y) + "\n"
                            file.write(result)
                            percent+=1
                
        file.close()
        print(str((percent/len(self.s2))*100))
             
    def levenshtein(self,str1,str2):
        size_x = len(str1) + 1
        size_y = len(str2) + 1
        matrix = np.zeros((size_x,size_y))
        for x in range(size_x):
            matrix [x,0] = x
        for y in range(size_y):
            matrix [0,y] = y
        
        for x in range(1,size_x):
            for y in range(1,size_y):
                if str1[x-1] == str2[y-1]:
                    matrix[x,y] = min(
                        matrix[x-1,y]+1,
                        matrix[x-1,y-1],
                        matrix[x,y-1]+1
                    )

                else:
                    matrix [x,y]=min(
                        matrix[x-1,y]+1,
                        matrix[x-1,y-1]+1,
                        matrix[x,y-1]+1
                    )
        return(matrix[size_x - 1, size_y - 1])

    
